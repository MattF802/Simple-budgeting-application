#Simple template made using chat gpt
import tkinter as tk
from tkinter import messagebox, simpledialog
import openpyxl
from openpyxl import load_workbook, Workbook
import matplotlib.pyplot as plt
from io import BytesIO
import tempfile
import os

# Define the budget dictionary
budget = {}

# Initialize the main application window
app = tk.Tk()
app.title("Budgeting Application")

def create_template_budget(filename):
    # Define the structure of the template budget
    categories = {
        'Income': 0,
        'Housing': 0,
        'Utilities': 0,
        'Food': 0,
        'Transportation': 0,
        'Healthcare': 0,
        'Entertainment': 0,
        'Savings': 0,
        'Miscellaneous': 0
    }

    # Create a new workbook and add the categories to it
    wb = Workbook()
    ws = wb.active
    ws.title = 'Budget'
    for row, (category, amount) in enumerate(categories.items(), start=1):
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=amount)

    # Save the template budget file
    wb.save(filename)

def load_budget(filename):
    if not os.path.isfile(filename):
        print(f"Budget file '{filename}' not found. Creating a new template budget file...")
        create_template_budget(filename)
    try:
        wb = load_workbook(filename)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        print("Creating a new template budget file...")
        create_template_budget(filename)
        wb = load_workbook(filename)
    return wb

# Usage example
filename = 'budget.xlsx'
wb = load_budget(filename)

# Function to update budget display
def update_budget_display():
    budget_display.delete(1.0, tk.END)
    for category, amount in budget.items():
        budget_display.insert(tk.END, f"{category}: ${amount}\n")

# Function to load budget data from Excel file
def load_budget():
    global budget_display
    filename = "budget.xlsx"
    try:
        wb = load_workbook(filename)
        ws = wb.active
        budget.clear()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                budget[row[0].lower()] = row[1]
        update_budget_display()
        messagebox.showinfo("Success", "Budget loaded from budget.xlsx")
    except FileNotFoundError:
        messagebox.showerror("Error", "Budget file not found.")

# Function to save budget data to Excel file
def save_budget():
    filename = "budget.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"

    headers = ['Category', 'Amount']
    ws.append(headers)

    for category, amount in budget.items():
        ws.append([category, amount])

    try:
        wb.save(filename)
        messagebox.showinfo("Success", "Budget saved to budget.xlsx")
    except PermissionError:
        messagebox.showerror("Error", "Could not save the budget. Please make sure the file is not open.")

def add_to_budget():
    category = category_entry.get().lower()
    amount_str = amount_entry.get()
    try:
        amount = float(amount_str)  # Convert amount to float
    except ValueError:
        messagebox.showerror("Error", "Please enter a valid amount.")
        return
    
    if category and amount:
        budget[category] = amount
        update_budget_display()
        category_entry.delete(0, tk.END)
        amount_entry.delete(0, tk.END)
    else:
        messagebox.showerror("Error", "Please enter both category and amount.")

def delete_category():
    category = simpledialog.askstring("Delete Category", "Enter category to delete:")
    if category:
        category = category.lower()
        if category in budget:
            del budget[category]
            update_budget_display()
        else:
            messagebox.showerror("Error", f"Category '{category}' not found.")

def update_category():
    old_category = simpledialog.askstring("Update Category", "Enter category to update:")
    if old_category:
        old_category = old_category.lower()
        if old_category in budget:
            new_amount = simpledialog.askfloat("Update Category", "Enter new amount:")
            if new_amount is not None:
                budget[old_category] = new_amount
                update_budget_display()
        else:
            messagebox.showerror("Error", f"Category '{old_category}' not found.")

def calculate_remaining():
    salary_str = salary_entry.get()
    if salary_str:
        try:
            salary = float(salary_str)  # Get the salary input
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid salary.")
            return
    else:
        messagebox.showerror("Error", "Please enter your monthly salary.")
        return

    total_expenses = sum(budget.values())
    remaining = salary - total_expenses
    remaining_label.config(text=f"Remaining budget: ${remaining}")
    
    # Update the budget with the remaining amount
    budget["Remaining"] = remaining
    update_budget_display()


import tempfile

def generate_pie_chart():
    labels = budget.keys()
    sizes = budget.values()
    plt.figure(figsize=(6, 6))
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
    plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle

    # Save the pie chart image to a temporary file
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
        plt.savefig(temp_file.name, format='png')
        temp_file.close()

        # Add the image file to the Excel workbook
        chart_sheet = wb.create_sheet("Pie Chart")
        img = openpyxl.drawing.image.Image(temp_file.name)
        chart_sheet.add_image(img, 'A1')
        chart_sheet.sheet_properties.tabColor = "00FF00"  # Set sheet tab color

    # Clean up the temporary file
    os.unlink(temp_file.name)


# Labels and Entries for user input
salary_label = tk.Label(app, text="Enter your monthly salary:")
salary_label.grid(row=0, column=0, padx=10, pady=5)
salary_entry = tk.Entry(app)
salary_entry.grid(row=0, column=1, padx=10, pady=5)
salary_entry.focus()  # Set focus to the salary entry widget

category_label = tk.Label(app, text="Category:")
category_label.grid(row=1, column=0, padx=10, pady=5)
category_entry = tk.Entry(app)
category_entry.grid(row=1, column=1, padx=10, pady=5)

amount_label = tk.Label(app, text="Amount:")
amount_label.grid(row=2, column=0, padx=10, pady=5)
amount_entry = tk.Entry(app)
amount_entry.grid(row=2, column=1, padx=10, pady=5)

add_button = tk.Button(app, text="Add to Budget", command=add_to_budget)
add_button.grid(row=3, column=0, columnspan=2, padx=10, pady=5)

# Displaying the budget
budget_display = tk.Text(app, height=10, width=30)
budget_display.grid(row=4, column=0, columnspan=2, padx=10, pady=5)

# Button to save budget
save_button = tk.Button(app, text="Save Budget", command=save_budget)
save_button.grid(row=5, column=0, columnspan=2, padx=10, pady=5)

# Label to display remaining budget
remaining_label = tk.Label(app, text="")
remaining_label.grid(row=6, column=0, columnspan=2, padx=10, pady=5)

# Calculate remaining budget
calculate_button = tk.Button(app, text="Calculate Remaining Budget", command=calculate_remaining)
calculate_button.grid(row=7, column=0, columnspan=2, padx=10, pady=5)

# Button to delete category
delete_button = tk.Button(app, text="Delete Category", command=delete_category)
delete_button.grid(row=8, column=0, columnspan=2, padx=10, pady=5)

# Button to update category
update_button = tk.Button(app, text="Update Category", command=update_category)
update_button.grid(row=9, column=0, columnspan=2, padx=10, pady=5)

# Load budget data from Excel file
load_budget()

# Run the application
app.mainloop()
